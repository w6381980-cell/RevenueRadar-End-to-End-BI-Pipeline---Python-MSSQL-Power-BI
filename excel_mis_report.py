# ============================================================
# excel_mis_report.py — RevenueRadar
# ============================================================
# YEH FILE KYA KARTI HAI:
#   Management Information System (MIS) report banata hai
#   Excel format mein — formatted, colored, professional.
#   Manager seedha yeh file email pe receive kar sakta hai.
#
# RUN: python 05_reports/excel_mis_report.py
# OUTPUT: 08_output/MIS_Report_April2025.xlsx
# ============================================================

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import sys, os
from datetime import datetime

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from config.config import get_engine, CLEANED_TABLE, OUTPUT_DIR, COMPANY_NAME

# ── Colors (hex) ─────────────────────────────────────────────
DARK_BLUE  = "1A3C6B"
MID_BLUE   = "2E6DA4"
LIGHT_BLUE = "D9E8F5"
ORANGE     = "D85A30"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
DARK_GRAY  = "444444"
GREEN      = "1D7A4A"
RED_COLOR  = "B22222"

def style_header(cell, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    """Header cell ko style karo."""
    cell.font      = Font(color=fg, size=size, bold=bold,
                          name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center",
                               vertical="center", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def generate_mis_report():
    print("📊 Excel MIS Report generate ho rahi hai...")

    # ── Data load karo ───────────────────────────────────────
    engine   = get_engine()
    df       = pd.read_sql(f"SELECT * FROM {CLEANED_TABLE}", engine)
    df["sale_date"] = pd.to_datetime(df["sale_date"])
    delivered = df[df["order_status"] == "Delivered"]

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename  = f"MIS_Report_{datetime.now().strftime('%B%Y')}.xlsx"
    filepath  = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Default sheet hatao

    # ════════════════════════════════════════════════════════
    # SHEET 1: Executive Summary
    # ════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 20

    # Title
    ws1.merge_cells("A1:C1")
    ws1["A1"] = f"REVENUERADAR — MIS REPORT | {datetime.now().strftime('%B %Y').upper()}"
    style_header(ws1["A1"], bg=DARK_BLUE, size=14)
    ws1.row_dimensions[1].height = 35

    ws1.merge_cells("A2:C2")
    ws1["A2"] = COMPANY_NAME
    ws1["A2"].font      = Font(color=DARK_GRAY, size=10, name="Calibri")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # KPI rows
    total_revenue = round(delivered["total_amount"].sum(), 2)
    total_orders  = len(delivered)
    avg_order_val = round(total_revenue / total_orders, 2) if total_orders else 0
    return_rate   = round(
        len(df[df["order_status"]=="Returned"]) / len(df) * 100, 2
    )

    # Monthly comparison
    monthly = (
        delivered.groupby(["year","month"])["total_amount"]
        .sum().reset_index().sort_values(["year","month"])
    )
    if len(monthly) >= 2:
        this_m  = monthly.iloc[-1]["total_amount"]
        last_m  = monthly.iloc[-2]["total_amount"]
        mom_pct = round(((this_m - last_m) / last_m) * 100, 2)
    else:
        mom_pct = 0

    kpi_data = [
        ("METRIC",            "VALUE",          "vs LAST MONTH"),
        ("Total Revenue",     f"₹{total_revenue:,.0f}", f"{mom_pct:+.1f}%"),
        ("Total Orders",      f"{total_orders:,}",    "—"),
        ("Avg Order Value",   f"₹{avg_order_val:,.0f}", "—"),
        ("Return Rate",       f"{return_rate}%",       "—"),
        ("Top Category",      delivered.groupby("category")["total_amount"].sum().idxmax(), "—"),
        ("Top Region",        delivered.groupby("region")["total_amount"].sum().idxmax(),   "—"),
    ]

    for row_num, (metric, value, change) in enumerate(kpi_data, start=4):
        ws1.row_dimensions[row_num].height = 24
        cells = [ws1.cell(row=row_num, column=c) for c in range(1, 4)]

        if row_num == 4:  # Header row
            cells[0].value = metric
            cells[1].value = value
            cells[2].value = change
            for c in cells:
                style_header(c, bg=MID_BLUE, size=10)
        else:
            cells[0].value = metric
            cells[1].value = value
            cells[2].value = change
            bg = LIGHT_GRAY if row_num % 2 == 0 else WHITE
            for c in cells:
                c.fill      = PatternFill("solid", fgColor=bg)
                c.font      = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
                c.border    = thin_border()

            # MoM color: green positive, red negative
            if "+" in str(change):
                cells[2].font = Font(color=GREEN, bold=True,
                                     size=10, name="Calibri")
            elif "-" in str(change) and "%" in str(change):
                cells[2].font = Font(color=RED_COLOR, bold=True,
                                     size=10, name="Calibri")

    print("   ✅ Sheet 1: Executive Summary done")

    # ════════════════════════════════════════════════════════
    # SHEET 2: Monthly Revenue Trend
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Monthly Trend")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "MONTHLY REVENUE TREND"
    style_header(ws2["A1"], bg=DARK_BLUE, size=13)
    ws2.row_dimensions[1].height = 30

    headers = ["Year", "Month", "Month Name", "Revenue (₹)", "Orders"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        style_header(c, bg=MID_BLUE, size=10)
        ws2.column_dimensions[get_column_letter(col)].width = 16

    monthly_full = (
        delivered.groupby(["year","month","month_name"])
        .agg(revenue=("total_amount","sum"), orders=("order_id","count"))
        .reset_index()
        .sort_values(["year","month"])
    )

    for i, row in enumerate(monthly_full.itertuples(), start=3):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data = [row.year, row.month, row.month_name,
                round(row.revenue, 2), row.orders]
        for col, val in enumerate(data, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center",
                                    vertical="center")
            c.border    = thin_border()

    print("   ✅ Sheet 2: Monthly Trend done")

    # ════════════════════════════════════════════════════════
    # SHEET 3: Category Analysis
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Category Analysis")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "CATEGORY-WISE PERFORMANCE"
    style_header(ws3["A1"], bg=DARK_BLUE, size=13)
    ws3.row_dimensions[1].height = 30

    cat_headers = ["Category", "Revenue (₹)", "Orders", "Revenue Share %"]
    for col, h in enumerate(cat_headers, 1):
        c = ws3.cell(row=2, column=col, value=h)
        style_header(c, bg=MID_BLUE, size=10)
        ws3.column_dimensions[get_column_letter(col)].width = 22

    cat_data = (
        delivered.groupby("category")
        .agg(revenue=("total_amount","sum"), orders=("order_id","count"))
        .reset_index()
        .sort_values("revenue", ascending=False)
    )
    total_rev = cat_data["revenue"].sum()
    cat_data["share"] = (cat_data["revenue"] / total_rev * 100).round(2)

    for i, row in enumerate(cat_data.itertuples(), start=3):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data = [row.category, round(row.revenue, 2),
                row.orders, f"{row.share}%"]
        for col, val in enumerate(data, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center",
                                    vertical="center")
            c.border    = thin_border()

    print("   ✅ Sheet 3: Category Analysis done")

    # ════════════════════════════════════════════════════════
    # SHEET 4: Raw Data (for reference)
    # ════════════════════════════════════════════════════════
    ws4    = wb.create_sheet("Data")
    ws4.sheet_view.showGridLines = True
    cols   = ["order_id","product_name","category","total_amount",
              "region","city","sale_date","order_status","payment_method"]
    sample = delivered[cols].head(5000)

    for col_num, col_name in enumerate(cols, 1):
        c = ws4.cell(row=1, column=col_num, value=col_name.upper())
        style_header(c, bg=DARK_BLUE, size=10)
        ws4.column_dimensions[get_column_letter(col_num)].width = 18

    for r, row in enumerate(sample.itertuples(index=False), start=2):
        for c, val in enumerate(row, 1):
            cell = ws4.cell(row=r, column=c, value=str(val))
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center")

    print("   ✅ Sheet 4: Raw data done")

    # ── Save karo ────────────────────────────────────────────
    wb.save(filepath)
    print(f"\n✅ Excel MIS Report saved: {filepath}")
    return filepath


if __name__ == "__main__":
    print("🚀 RevenueRadar — Excel MIS Report Generator")
    print("=" * 55)
    path = generate_mis_report()
    print(f"\n📁 Open karo: {path}")
    print("➡️  Agla step: python 05_reports/impact_report.py")